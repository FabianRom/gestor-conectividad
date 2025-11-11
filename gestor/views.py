from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import (
    Escuela, Region, Predio, Distrito, TipoEstablecimiento, Categoria, 
    ServicioConectividad, PisoTecnologico, Dependencia, Ambito, Turno, 
    Ciudad, PlanPiso, EstadoConectividad, 
    ProveedorInternet, ProveedorPisoTecnologico, # Proveedores ya existentes
    TipoPisoTecnologico, MetodoSolicitud, # Modelos agregados para Carga Masiva
)
from django.db.models import Q, Count
# Importaciones necesarias al inicio del archivo excel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from django.db.models import Count, F, ExpressionWrapper, DecimalField, Sum, Case, When, Value, BooleanField
from django.db.models.functions import Cast # Asegúrate de que Cast esté importado
from django.core.paginator import Paginator




# --- Importaciones para la gestión de archivos (CSV) ---
import csv
import io
from datetime import datetime


# =========================================================================
# --- MENU
# Vistas necesarias para el menú:
# =========================================================================



# =========================================================================
# --- Vistas Generales ---
# =========================================================================

def home(request):
    """Renderiza la página de inicio."""
    return render(request, 'gestor/home.html')

def lista_escuelas(request):
    """
    Lista de escuelas con filtros por región, CUE y predio.
    """
    region_id = request.GET.get('region', '')
    cue = request.GET.get('cue', '')
    predio_id = request.GET.get('predio', '')

    escuelas = Escuela.objects.all().select_related(
        'region', 'predio', 'categoria'
    )

    if region_id:
        escuelas = escuelas.filter(region_id=region_id)

    if cue:
        escuelas = escuelas.filter(cue__icontains=cue)

    if predio_id:
        escuelas = escuelas.filter(predio_id=predio_id)

    regiones = Region.objects.all().order_by('nombre')
    predios = Predio.objects.all().order_by('numero_predio')

    mensaje_sin_resultados = None
    if not escuelas.exists():
        mensaje_sin_resultados = "No se encontraron escuelas que coincidan con la búsqueda."

    context = {
        'escuelas': escuelas,
        'regiones': regiones,
        'predios': predios,
        'region_seleccionada': region_id,
        'cue_seleccionado': cue,
        'predio_seleccionado': predio_id,
        'mensaje_sin_resultados': mensaje_sin_resultados,
    }

    return render(request, 'gestor/lista_escuelas.html', context)


def detalle_escuela(request, cue):
    """
    Muestra el detalle de una escuela específica, incluyendo sus servicios y pisos tecnológicos.
    """
    escuela = get_object_or_404(Escuela, cue=cue)
    servicios = ServicioConectividad.objects.filter(escuela=escuela)
    pisos_tecnologicos = PisoTecnologico.objects.filter(escuela=escuela)
    
    # NUEVO: obtener otras escuelas con el mismo predio
    otras_escuelas_predio = Escuela.objects.filter(
        predio=escuela.predio
    ).exclude(cue=escuela.cue)
    
    context = {
        'escuela': escuela,
        'servicios': servicios,
        'pisos_tecnologicos': pisos_tecnologicos,
        'otras_escuelas_predio': otras_escuelas_predio,
    }
    return render(request, 'gestor/detalle_escuela.html', context)


#Api para mapa escuelas colores 
def api_escuela(request, cue):
    escuela = get_object_or_404(Escuela, cue=cue)

    servicio = ServicioConectividad.objects.filter(escuela=escuela).first()
    piso = PisoTecnologico.objects.filter(escuela=escuela).first()

    data = {
        "nombre": escuela.nombre,
        "cue": escuela.cue,
        "categoria": escuela.categoria.nombre if escuela.categoria else "Sin dato",

        # Plan de enlace (Estado de Conectividad)
        "plan_enlace": servicio.estado_conectividad.nombre if servicio and servicio.estado_conectividad else "Sin dato",

        # Proveedor internet
        "proveedor_internet": servicio.proveedor.nombre if servicio and servicio.proveedor else "Sin dato",

        # Plan de Piso
        "plan_piso": piso.plan_piso.nombre if piso and piso.plan_piso else "Sin dato",

        # Tipo de Piso Instalado
        "tipo_piso": piso.tipo_piso_instalado.nombre if piso and piso.tipo_piso_instalado else "Sin dato",
    }

    return JsonResponse(data)



# =========================================================================
# --- Vistas de Búsqueda Avanzada ---
# =========================================================================

def busqueda(request):
    """
    Renderiza la página de búsqueda avanzada con TODAS las opciones de filtro de catálogo.
    """
    context = {
        'distritos': Distrito.objects.all().order_by('nombre'),
        'dependencias': Dependencia.objects.all().order_by('nombre'),
        'estados_conectividad': EstadoConectividad.objects.all().order_by('nombre'),
        'planes_piso': PlanPiso.objects.all().order_by('nombre'),
        'regiones': Region.objects.all().order_by('nombre'),
        'predios': Predio.objects.all().order_by('numero_predio'), 
        'ciudades': Ciudad.objects.all().order_by('nombre'),
        'ambitos': Ambito.objects.all().order_by('nombre'),
        'turnos': Turno.objects.all().order_by('nombre'),
        'categorias': Categoria.objects.all().order_by('nombre'),
        'tipos_establecimiento': TipoEstablecimiento.objects.all().order_by('nombre'),
        'proveedores_internet': ProveedorInternet.objects.all().order_by('nombre'), 
        'proveedores_piso': ProveedorPisoTecnologico.objects.all().order_by('nombre'), 
    }
    return render(request, 'gestor/busqueda.html', context)

def resultados_busqueda(request):
    """
    Vista para mostrar los resultados de la búsqueda de escuelas según filtros.
    Se han corregido los nombres de las variables de GET para coincidir con el HTML.
    """
    # --- Filtros básicos y geográficos ---
    cue = request.GET.get('cue', '').strip()
    nombre = request.GET.get('nombre', '').strip()
    predio_numero = request.GET.get('predio_numero', '').strip()
    region_id = request.GET.get('region', '')
    distrito_id = request.GET.get('distrito', '')
    ciudad_id = request.GET.get('ciudad', '') 

    # --- Atributos Institucionales ---
    ambito_id = request.GET.get('ambito', '')
    dependencia_id = request.GET.get('dependencia', '')
    turno_id = request.GET.get('turno', '')
    categoria_id = request.GET.get('categoria', '')
    tipo_establecimiento_id = request.GET.get('tipo_establecimiento', '')

    # --- Conectividad y Tecnología (Variables corregidas para coincidir con el HTML) ---
    proveedor_internet_id = request.GET.get('proveedor_internet', '')
    proveedor_piso_id = request.GET.get('proveedor_piso', '')

    # 🚨 CORRECCIÓN 1: El HTML usa 'conectividad' para el Estado de Conectividad.
    estado_conectividad_id = request.GET.get('conectividad', '') 
    
    # 🚨 CORRECCIÓN 2: El HTML usa 'programa_conectividad' para el Plan Piso.
    plan_piso_id = request.GET.get('programa_conectividad', '') 

    # --- Opciones binarias ---
    tiene_internet = request.GET.get('tiene_internet', '')
    tiene_piso_tecnologico = request.GET.get('tiene_piso_tecnologico', '')
    
    # --- Filtros de Año (Asumiendo que se quieren usar) ---
    ano_conectado = request.GET.get('ano_conectado', '')
    ano_finalizacion_piso = request.GET.get('ano_finalizacion_piso', '')

    # --- Queryset inicial ---
    queryset = Escuela.objects.all().select_related(
        'region', 'distrito', 'ciudad', 'predio', 'ambito', 'dependencia', 
        'turno', 'categoria', 'tipo_establecimiento'
    )

    # --- Aplicar filtros básicos ---
    if cue:
        queryset = queryset.filter(cue__icontains=cue)
    if nombre:
        queryset = queryset.filter(nombre__icontains=nombre)
    if predio_numero:
        queryset = queryset.filter(predio__numero_predio__icontains=predio_numero)
    if region_id:
        queryset = queryset.filter(region_id=region_id)
    if distrito_id:
        queryset = queryset.filter(distrito_id=distrito_id)
    if ciudad_id:
        queryset = queryset.filter(ciudad_id=ciudad_id) 

    # --- Aplicar filtros institucionales ---
    if ambito_id:
        queryset = queryset.filter(ambito_id=ambito_id)
    if dependencia_id:
        queryset = queryset.filter(dependencia_id=dependencia_id)
    if turno_id:
        queryset = queryset.filter(turno_id=turno_id)
    if categoria_id:
        queryset = queryset.filter(categoria_id=categoria_id)
    if tipo_establecimiento_id:
        queryset = queryset.filter(tipo_establecimiento_id=tipo_establecimiento_id)

    # --- Aplicar filtros de conectividad y tecnología (CORREGIDOS Y RELACIONADOS) ---

    # Filtro por Estado de Conectividad (Relación: ServicioConectividad)
    if estado_conectividad_id:
        queryset = queryset.filter(
            servicioconectividad__estado_conectividad_id=estado_conectividad_id
        )
    
    # Filtro por Plan Piso Tecnológico (Relación: PisoTecnologico)
    if plan_piso_id:
        queryset = queryset.filter(
            pisotecnologico__plan_piso_id=plan_piso_id
        )

    # Filtro por Proveedor de Internet
    if proveedor_internet_id:
        queryset = queryset.filter(servicioconectividad__proveedor_id=proveedor_internet_id)
        
    # Filtro por Proveedor Piso Tecnológico
    if proveedor_piso_id:
        queryset = queryset.filter(pisotecnologico__proveedor_id=proveedor_piso_id)
        
    # Filtro por Año de Conexión (Asumiendo que se usa el año de instalación o mejora)
    if ano_conectado:
        # Busca escuelas donde la fecha de instalación o mejora tenga el año ingresado
        queryset = queryset.filter(
            servicioconectividad__fecha_instalacion__year=ano_conectado
        ) | queryset.filter(
            servicioconectividad__fecha_mejora__year=ano_conectado
        )

    # Filtro por Año de Finalización de Piso Tecnológico
    if ano_finalizacion_piso:
        queryset = queryset.filter(pisotecnologico__fecha_terminado__year=ano_finalizacion_piso)


    # Opciones binarias (Estas deberían estar en el modelo Escuela)
    if tiene_internet == 'si':
        queryset = queryset.filter(tiene_internet=True)
    elif tiene_internet == 'no':
        queryset = queryset.filter(tiene_internet=False)
    
    if tiene_piso_tecnologico == 'si':
        queryset = queryset.filter(tiene_piso_tecnologico=True)
    elif tiene_piso_tecnologico == 'no':
        queryset = queryset.filter(tiene_piso_tecnologico=False)

    # Es CRUCIAL usar .distinct() después de filtrar por relaciones inversas
    # para evitar que una escuela aparezca varias veces en los resultados.
    queryset = queryset.distinct()

    # --- Paginación ---
    paginator = Paginator(queryset, 20) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    encabezado_resultado = f"Resultados de la búsqueda: {queryset.count()} escuela(s) encontrada(s)"

    context = {
        'page_obj': page_obj,
        'resultados': page_obj.object_list,
        'encabezado_resultado': encabezado_resultado,
        'filtros': request.GET, # Para que la plantilla pueda mantener los valores seleccionados
    }
    
    # Asegúrate de que las variables de contexto para los filtros (regiones, etc.)
    # se pasen si se necesita renderizar el formulario nuevamente. 
    # (Asumo que esta lógica está en otra vista o en resultados_busqueda.html)

    return render(request, 'gestor/resultados_busqueda.html', context)

# =========================================================================
# --- Vistas de Gestión de Datos (Carga Masiva) ---
# =========================================================================

def carga_descarga_view(request):
    """Renderiza el template para la gestión de importación y exportación."""
    return render(request, 'gestor/carga_descarga.html')

# Función auxiliar para obtener o crear objetos de catálogo
def get_or_create_related(model, value):
    """Busca un objeto de catálogo por nombre o lo crea si no existe."""
    if not value or not value.strip():
        return None
    # Usamos strip() para limpiar espacios en blanco antes de buscar/crear
    obj, created = model.objects.get_or_create(nombre=value.strip())
    return obj

@transaction.atomic
def importar_datos(request):
    """Procesa el archivo CSV subido para crear/actualizar datos."""
    if request.method != 'POST':
        messages.error(request, 'Error: Se esperaba una solicitud POST.')
        return redirect('carga_descarga_url')
        
    if 'csv_file' not in request.FILES:
        messages.error(request, 'Error: No se ha adjuntado ningún archivo.')
        return redirect('carga_descarga_url')

    csv_file = request.FILES['csv_file']
    
    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'Error: El archivo debe ser un CSV.')
        return redirect('carga_descarga_url')

    try:
        # Decodificar el archivo y manejar el formato de entrada
        data_set = csv_file.read().decode('utf-8')
        io_string = io.StringIO(data_set)
        
        # Saltamos el encabezado (asumimos que la primera fila es el header)
        next(io_string) 

        imported_count = 0
        updated_count = 0
        errors = []

        reader = csv.reader(io_string, delimiter=',')
        
        # Definición de índices de columna
        CUE_INDEX = 0
        CLAVE_PROVINCIAL_INDEX = 1
        NOMBRE_INDEX = 2
        DIRECCION_INDEX = 3
        MATRICULA_INDEX = 4
        LATITUD_INDEX = 5
        LONGITUD_INDEX = 6
        
        REGION_INDEX = 7
        DISTRITO_INDEX = 8
        CIUDAD_INDEX = 9
        AMBITO_INDEX = 10
        DEPENDENCIA_INDEX = 11
        TURNO_INDEX = 12
        CATEGORIA_INDEX = 13
        TIPO_ESTABLECIMIENTO_INDEX = 14
        NUMERO_PREDIO_INDEX = 15

        INTERNET_TIENE_INDEX = 16
        INTERNET_PROV_INDEX = 17
        INTERNET_VELOCIDAD_INDEX = 18
        INTERNET_ESTADO_INDEX = 19
        INTERNET_FECHA_INDEX = 20
        INTERNET_METODO_INDEX = 21
        INTERNET_OBS_INDEX = 22

        PISO_TIENE_INDEX = 23
        PISO_PROV_INDEX = 24
        PISO_PLAN_INDEX = 25
        PISO_TIPO_INDEX = 26
        PISO_FECHA_INDEX = 27
        PISO_MEJORA_INDEX = 28
        PISO_OBS_INDEX = 29
        
        for i, row in enumerate(reader):
            # i es el índice de la fila de datos (después del encabezado)
            
            if not row or not row[CUE_INDEX].strip(): continue 
            
            cue = row[CUE_INDEX].strip()
            
            try:
                # --- Validaciones y Conversiones ---
                matricula = int(row[MATRICULA_INDEX] or 0)
                tiene_internet = row[INTERNET_TIENE_INDEX].strip().upper() in ('SÍ', 'SI', 'TRUE')
                tiene_piso = row[PISO_TIENE_INDEX].strip().upper() in ('SÍ', 'SI', 'TRUE')
                
                # --- Predio (Obligatorio) ---
                predio_num = row[NUMERO_PREDIO_INDEX].strip()
                if not predio_num:
                    raise ValueError("El número de predio no puede estar vacío.")
                predio, _ = Predio.objects.get_or_create(numero_predio=predio_num)
                
                # --- Escuela: Buscar o crear objetos relacionados (Catálogos) ---
                region = get_or_create_related(Region, row[REGION_INDEX])
                distrito = get_or_create_related(Distrito, row[DISTRITO_INDEX])
                ciudad = get_or_create_related(Ciudad, row[CIUDAD_INDEX])
                ambito = get_or_create_related(Ambito, row[AMBITO_INDEX])
                dependencia = get_or_create_related(Dependencia, row[DEPENDENCIA_INDEX])
                turno = get_or_create_related(Turno, row[TURNO_INDEX])
                categoria = get_or_create_related(Categoria, row[CATEGORIA_INDEX])
                tipo_establecimiento = get_or_create_related(TipoEstablecimiento, row[TIPO_ESTABLECIMIENTO_INDEX])
                
                # Crear o actualizar Escuela
                escuela_data = {
                    'clave_provincial': row[CLAVE_PROVINCIAL_INDEX].strip() or None,
                    'nombre': row[NOMBRE_INDEX].strip(),
                    'direccion': row[DIRECCION_INDEX].strip(),
                    'matricula': matricula,
                    'latitud': row[LATITUD_INDEX] or None,
                    'longitud': row[LONGITUD_INDEX] or None,
                    'tiene_internet': tiene_internet,
                    'tiene_piso_tecnologico': tiene_piso,
                    'predio': predio,
                    'region': region, 'distrito': distrito, 'ciudad': ciudad, 
                    'ambito': ambito, 'dependencia': dependencia, 'turno': turno, 
                    'categoria': categoria, 'tipo_establecimiento': tipo_establecimiento
                }
                
                escuela, created = Escuela.objects.update_or_create(
                    cue=cue,
                    defaults=escuela_data
                )
                
                if created:
                    imported_count += 1
                else:
                    updated_count += 1

                # --- ServicioConectividad ---
                if escuela.tiene_internet:
                    prov_int = get_or_create_related(ProveedorInternet, row[INTERNET_PROV_INDEX])
                    estado_con = get_or_create_related(EstadoConectividad, row[INTERNET_ESTADO_INDEX])
                    metodo_sol = get_or_create_related(MetodoSolicitud, row[INTERNET_METODO_INDEX])
                    
                    fecha_inst_str = row[INTERNET_FECHA_INDEX].strip()
                    fecha_inst = datetime.strptime(fecha_inst_str, '%Y-%m-%d').date() if fecha_inst_str else None
                    
                    ServicioConectividad.objects.update_or_create(
                        escuela=escuela,
                        defaults={
                            'proveedor': prov_int,
                            'velocidad_mbps': int(row[INTERNET_VELOCIDAD_INDEX] or 0),
                            'estado_conectividad': estado_con,
                            'fecha_instalacion': fecha_inst,
                            'metodo_solicitud': metodo_sol,
                            'observaciones': row[INTERNET_OBS_INDEX].strip(),
                        }
                    )
                # Si no tiene internet, eliminar el registro de servicio si existe
                elif ServicioConectividad.objects.filter(escuela=escuela).exists():
                     ServicioConectividad.objects.filter(escuela=escuela).delete()


                # --- PisoTecnologico ---
                if escuela.tiene_piso_tecnologico:
                    prov_piso = get_or_create_related(ProveedorPisoTecnologico, row[PISO_PROV_INDEX])
                    plan_piso = get_or_create_related(PlanPiso, row[PISO_PLAN_INDEX])
                    tipo_piso = get_or_create_related(TipoPisoTecnologico, row[PISO_TIPO_INDEX])
                    
                    fecha_term_str = row[PISO_FECHA_INDEX].strip()
                    fecha_term = datetime.strptime(fecha_term_str, '%Y-%m-%d').date() if fecha_term_str else None

                    PisoTecnologico.objects.update_or_create(
                        escuela=escuela,
                        defaults={
                            'proveedor': prov_piso,
                            'plan_piso': plan_piso,
                            'tipo_piso_instalado': tipo_piso,
                            'fecha_terminado': fecha_term,
                            'tipo_mejora': row[PISO_MEJORA_INDEX].strip(),
                            'observaciones': row[PISO_OBS_INDEX].strip(),
                        }
                    )
                # Si no tiene piso, eliminar el registro de piso si existe
                elif PisoTecnologico.objects.filter(escuela=escuela).exists():
                    PisoTecnologico.objects.filter(escuela=escuela).delete()

            except Exception as e:
                # Recolectar errores y continuar con la siguiente fila
                errors.append(f"Fila {i+2} (CUE: {cue}): Error al procesar - {e}")
                # Imprimir errores en la consola del servidor para depuración
                print(f"Error de Importación Fila {i+2}: {e}") 

        # Mensajes al finalizar la transacción atómica
        if errors:
            messages.error(request, f'Carga masiva finalizada con {len(errors)} errores. Total: {imported_count} creadas, {updated_count} actualizadas.')
        else:
            messages.success(request, f'Carga masiva exitosa: {imported_count} escuelas creadas, {updated_count} escuelas actualizadas.')

    except Exception as general_e:
        # Manejar errores que ocurren fuera del bucle de filas (ej. archivo malo o codificación)
        messages.error(request, f'Error general durante la carga masiva: {general_e}')
        # Aseguramos que la transacción se revierta si ocurre un error general
        transaction.set_rollback(True)
        
    return redirect('carga_descarga_url')

def exportar_datos(request):
    """Exporta todas las escuelas y sus datos relacionados a un solo archivo CSV."""
    response = HttpResponse(content_type='text/csv')
    # Añade la fecha y hora al nombre del archivo
    filename = "escuelas_full_export_{}.csv".format(datetime.now().strftime('%Y%m%d_%H%M'))
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    
    # ---------------------------------------------------------------------
    # ESTRUCTURA DEL ENCABEZADO (Debe coincidir EXACTAMENTE con el orden en importación)
    # ---------------------------------------------------------------------
    header = [
        'CUE', 'Clave_Provincial', 'Nombre', 'Direccion', 'Matricula', 
        'Latitud', 'Longitud', 
        
        # Relaciones Geográficas/Institucionales (Catálogos)
        'Region', 'Distrito', 'Ciudad', 'Ambito', 'Dependencia', 'Turno', 'Categoria', 
        'Tipo_Establecimiento', 'Numero_Predio',
        
        # Datos de Conectividad (ServicioConectividad)
        'Internet_Tiene (Sí/No)', 'Internet_Proveedor', 'Internet_Velocidad_Mbps', 
        'Internet_Estado_Conectividad', 'Internet_Fecha_Instalacion (AAAA-MM-DD)', 
        'Internet_Metodo_Solicitud', 'Internet_Observaciones',
        
        # Datos de Piso Tecnológico (PisoTecnologico)
        'Piso_Tiene (Sí/No)', 'Piso_Proveedor', 'Piso_Plan', 
        'Piso_Tipo_Instalado', 'Piso_Fecha_Terminado (AAAA-MM-DD)', 'Piso_Tipo_Mejora', 
        'Piso_Observaciones',
    ]
    writer.writerow(header)

    # Optimizamos la consulta cargando todas las relaciones en una sola vez (select_related)
    escuelas = Escuela.objects.all().select_related(
        'region', 'distrito', 'ciudad', 'ambito', 'dependencia', 'turno', 
        'categoria', 'tipo_establecimiento', 'predio'
    )
    
    for escuela in escuelas:
        # Obtenemos servicios relacionados
        servicio = ServicioConectividad.objects.filter(escuela=escuela).first()
        piso = PisoTecnologico.objects.filter(escuela=escuela).first()

        # Construimos la fila
        row = [
            escuela.cue,
            escuela.clave_provincial or '',
            escuela.nombre,
            escuela.direccion,
            escuela.matricula,
            escuela.latitud or '', 
            escuela.longitud or '',
            
            # Catálogos
            escuela.region.nombre if escuela.region else '',
            escuela.distrito.nombre if escuela.distrito else '',
            escuela.ciudad.nombre if escuela.ciudad else '',
            escuela.ambito.nombre if escuela.ambito else '',
            escuela.dependencia.nombre if escuela.dependencia else '',
            escuela.turno.nombre if escuela.turno else '',
            escuela.categoria.nombre if escuela.categoria else '',
            escuela.tipo_establecimiento.nombre if escuela.tipo_establecimiento else '',
            escuela.predio.numero_predio if escuela.predio else '',
            
            # Conectividad
            'Sí' if escuela.tiene_internet else 'No',
            servicio.proveedor.nombre if servicio and servicio.proveedor else '',
            servicio.velocidad_mbps if servicio else 0,
            servicio.estado_conectividad.nombre if servicio and servicio.estado_conectividad else '',
            servicio.fecha_instalacion.isoformat() if servicio and servicio.fecha_instalacion else '',
            servicio.metodo_solicitud.nombre if servicio and servicio.metodo_solicitud else '',
            servicio.observaciones.replace('\n', ' ') if servicio and servicio.observaciones else '',
            
            # Piso Tecnológico
            'Sí' if escuela.tiene_piso_tecnologico else 'No',
            piso.proveedor.nombre if piso and piso.proveedor else '',
            piso.plan_piso.nombre if piso and piso.plan_piso else '',
            piso.tipo_piso_instalado.nombre if piso and piso.tipo_piso_instalado else '',
            piso.fecha_terminado.isoformat() if piso and piso.fecha_terminado else '',
            piso.tipo_mejora if piso else '',
            piso.observaciones.replace('\n', ' ') if piso and piso.observaciones else '',
        ]
        writer.writerow(row)

    return response

def descargar_plantilla(request):
    """Descarga un archivo CSV vacío con los encabezados esperados para la importación."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="plantilla_carga_masiva.csv"'
    
    writer = csv.writer(response)
    
    # Encabezados - deben coincidir EXACTAMENTE con las funciones de importación/exportación
    header = [
        'CUE', 'Clave_Provincial', 'Nombre', 'Direccion', 'Matricula', 
        'Latitud', 'Longitud', 
        
        'Region', 'Distrito', 'Ciudad', 'Ambito', 'Dependencia', 'Turno', 'Categoria', 
        'Tipo_Establecimiento', 'Numero_Predio',
        
        # Datos de Conectividad
        'Internet_Tiene (Sí/No)', 'Internet_Proveedor', 'Internet_Velocidad_Mbps', 
        'Internet_Estado_Conectividad', 'Internet_Fecha_Instalacion (AAAA-MM-DD)', 
        'Internet_Metodo_Solicitud', 'Internet_Observaciones',
        
        # Datos de Piso Tecnológico
        'Piso_Tiene (Sí/No)', 'Piso_Proveedor', 'Piso_Plan', 
        'Piso_Tipo_Instalado', 'Piso_Fecha_Terminado (AAAA-MM-DD)', 'Piso_Tipo_Mejora', 
        'Piso_Observaciones',
    ]
    writer.writerow(header)
    
    # Fila de ejemplo para guiar al usuario
    writer.writerow([
        '012345678', 'CP001', 'Escuela Modelo', 'Av. Siempre Viva 123', '500', 
        '-34.6037', '-58.3816', 
        'Región 1', 'Distrito Central', 'Capital', 'Urbano', 'Nacional', 'Mañana', 'Nivel Primario', 
        'Pública', 'P-001',
        
        # Conectividad
        'Sí', 'Telecom', '100', 'Activo', '2023-01-15', 'Presencial', 'Buena velocidad',
        
        # Piso Tecnológico
        'No', '', '', '', '', '', ''
    ])

    return response


# =========================================================================
# --- Resto de Vistas de Reportes y Generación de Excel ---
# =========================================================================

from django.shortcuts import render
# Asegúrate de importar el modelo Escuela
from .models import Escuela 

# Usa la función con el nombre que tienes: reporte_internet
def reporte_internet(request):
    # 1. Obtener los conteos
    con_internet = Escuela.objects.filter(tiene_internet=True).count()
    # Asumimos que el resto es "sin internet"
    total_escuelas = Escuela.objects.count()
    sin_internet = total_escuelas - con_internet 

    # 2. CALCULAR LA TASA DE COBERTURA
    if total_escuelas > 0:
        # Usamos float() para garantizar una división decimal
        tasa_cobertura = (float(con_internet) / total_escuelas) * 100
        # Redondeamos a un decimal para presentación
        tasa_cobertura_formateada = round(tasa_cobertura, 1)
    else:
        tasa_cobertura_formateada = 0.0

    # 3. Definir el contexto
    contexto = {
        'titulo_reporte': 'Reporte de Conectividad a Internet',
        'con_internet': con_internet,
        'sin_internet': sin_internet,
        'tasa_cobertura': tasa_cobertura_formateada, # ¡Nueva variable para el porcentaje!
        'total_escuelas': total_escuelas, # Total para referencia
    }

    # Asegúrate de que renderice a reporte_internet.html
    return render(request, 'gestor/reporte_internet.html', contexto)
###
# Función para el Reporte de Piso Tecnológico
def reporte_piso(request):
    # 1. Obtener los conteos
    con_piso = Escuela.objects.filter(tiene_piso_tecnologico=True).count()
    total_escuelas = Escuela.objects.count()
    sin_piso = total_escuelas - con_piso 

    # 2. CALCULAR LA TASA DE COBERTURA DE PISO TECNOLÓGICO
    if total_escuelas > 0:
        # Usamos float() para garantizar una división decimal
        tasa_cobertura = (float(con_piso) / total_escuelas) * 100
        # Redondeamos a un decimal para presentación
        tasa_cobertura_formateada = round(tasa_cobertura, 1)
    else:
        tasa_cobertura_formateada = 0.0

    # 3. Definir el contexto
    contexto = {
        'titulo_reporte': 'Reporte de Piso Tecnológico',
        'con_piso': con_piso,
        'sin_piso': sin_piso,
        'tasa_cobertura': tasa_cobertura_formateada, # Tasa de cobertura de Piso
        'total_escuelas': total_escuelas, 
    }

    # Asegúrate de que renderice a reporte_piso.html
    return render(request, 'gestor/reporte_piso.html', contexto)

######


def dashboard(request):
    # --- 1. CÁLCULOS GLOBALES ---
    total_escuelas = Escuela.objects.count()
    con_internet = Escuela.objects.filter(tiene_internet=True).count()
    sin_internet = total_escuelas - con_internet
    con_piso = Escuela.objects.filter(tiene_piso_tecnologico=True).count()
    sin_piso = total_escuelas - con_piso
    
    # Cálculo del porcentaje global
    if total_escuelas > 0:
        tasa_cobertura = (float(con_internet) / total_escuelas) * 100
        # Formatea a dos decimales
        porcentaje_conectividad = "{:.2f}".format(tasa_cobertura) 
    else:
        porcentaje_conectividad = "0.00"
        
    # --------------------------------------------------------------------------
    # --- 2. CÁLCULOS: CONECTIVIDAD POR PROGRAMA (PBA / PNCE) ---
    # --------------------------------------------------------------------------
    
    # 2.1. Conectadas por PNCE (SOLO 'PNCE')
    # Buscamos escuelas con internet cuyo estado de conectividad es EXACTAMENTE 'PNCE'
    conectadas_pnce = Escuela.objects.filter(
        tiene_internet=True,
        servicioconectividad__estado_conectividad__nombre__iexact='PNCE' 
    ).count()

    # 2.2. Conectadas por PBA (INCLUYENDO 'PBA - PNCE')
    # Creamos las dos condiciones Q:
    filtro_pba_solo = Q(servicioconectividad__estado_conectividad__nombre__iexact='PBA')
    filtro_pnce_pba = Q(servicioconectividad__estado_conectividad__nombre__iexact='PNCE - PBA')
    
    # Combinamos la condición de 'tiene_internet=True' con el OR (|) de los estados
    conectadas_pba = Escuela.objects.filter(
        Q(tiene_internet=True) & (filtro_pba_solo | filtro_pnce_pba)
    ).count()
    
    # --------------------------------------------------------------------------
    # --- 3. CÁLCULO PARA LA GRÁFICA DE CATEGORÍAS ---
    # --------------------------------------------------------------------------
    
    # Usamos la lógica eficiente de filter() + annotate(Count()) para la gráfica
    conectadas_por_categoria_queryset = Categoria.objects.filter(
        escuela__tiene_internet=True
    ).annotate(
        total_conectadas=Count('escuela')
    ).values('nombre', 'total_conectadas')

    # Convertimos el QuerySet a una lista de diccionarios
    conectadas_por_categoria = list(conectadas_por_categoria_queryset)
    
    # --------------------------------------------------------------------------
    # --- 4. CONTEXTO ---
    # --------------------------------------------------------------------------
    contexto = {
        'total_escuelas': total_escuelas,
        'con_internet': con_internet,
        'sin_internet': sin_internet,
        'con_piso': con_piso,
        'sin_piso': sin_piso,
        
        # Variable usada en el bloque de porcentaje de la plantilla
        'porcentaje_conectividad': porcentaje_conectividad, 
        
        # Nuevas variables de los programas
        'conectadas_pba': conectadas_pba,
        'conectadas_pnce': conectadas_pnce,
        
        # Variable que alimenta la gráfica (¡ahora definida!)
        'conectadas_por_categoria': conectadas_por_categoria, 
    }
    return render(request, 'gestor/dashboard.html', contexto)


def dashboard_data(request):
    """Devuelve datos en JSON para gráficos del dashboard."""
    total_escuelas = Escuela.objects.count()
    con_internet = Escuela.objects.filter(tiene_internet=True).count()
    sin_internet = total_escuelas - con_internet
    con_piso = Escuela.objects.filter(tiene_piso_tecnologico=True).count()
    sin_piso = total_escuelas - con_piso
    
    # 1. CÁLCULO DE LA TASA DE COBERTURA (Global)
    if total_escuelas > 0:
        tasa_cobertura = (float(con_internet) / total_escuelas) * 100
        porcentaje_conectividad = "{:.2f}".format(tasa_cobertura)
    else:
        porcentaje_conectividad = "0.00"

    # 2. ESCUELAS CONECTADAS POR CATEGORÍA (Para la gráfica de barras)
    # Usamos el filtro 'escuela__tiene_internet=True' y contamos
    conectadas_por_categoria_json = Categoria.objects.filter(
        escuela__tiene_internet=True
    ).annotate(
        total_conectadas=Count('escuela')
    ).values('nombre', 'total_conectadas')
    
    data = {
        'total_escuelas': total_escuelas,
        'con_internet': con_internet,
        'sin_internet': sin_internet,
        'con_piso': con_piso,
        'sin_piso': sin_piso,
        'porcentaje_conectividad': porcentaje_conectividad, # Incluimos la tasa en el JSON
        
        # Enviamos la data de las escuelas CONECTADAS por categoría
        'conectadas_por_categoria': list(conectadas_por_categoria_json), 
    }
    return JsonResponse(data)



# Genera el  excel  para detalles completo  escuela 
def generar_excel_escuela(request, cue):
    """Genera un archivo Excel detallado para una escuela específica."""
    escuela = get_object_or_404(Escuela, cue=cue)
    servicios = ServicioConectividad.objects.filter(escuela=escuela)
    pisos_tecnologicos = PisoTecnologico.objects.filter(escuela=escuela)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="escuela_{escuela.cue}.xlsx"'
    
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Reporte de Escuela"

    fuente_negrita = Font(bold=True)
    relleno_cabecera = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    borde_fino = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    row_num = 1

    # Bloque de Información General
    hoja.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
    hoja[f'A{row_num}'] = "INFORMACIÓN GENERAL"
    hoja[f'A{row_num}'].font = fuente_negrita
    hoja[f'A{row_num}'].fill = relleno_cabecera
    hoja[f'A{row_num}'].alignment = Alignment(horizontal='center')
    row_num += 1

    datos_escuela = [
        ("Nombre", escuela.nombre),
        ("CUE", escuela.cue),
        ("Clave Provincial", escuela.clave_provincial),
        ("Dirección", escuela.direccion),
        ("Matrícula", escuela.matricula),
        ("Dependencia", escuela.dependencia.nombre if escuela.dependencia else 'Sin datos'),
        ("Ámbito", escuela.ambito.nombre if escuela.ambito else 'Sin datos'),
        ("Turno", escuela.turno.nombre if escuela.turno else 'Sin datos'),
        ("Categoría", escuela.categoria.nombre if escuela.categoria else 'Sin datos'),
        ("Tipo de Establecimiento", escuela.tipo_establecimiento.nombre if escuela.tipo_establecimiento else 'Sin datos'),
        ("Región", escuela.region.nombre if escuela.region else 'Sin datos'),
        ("Distrito", escuela.distrito.nombre if escuela.distrito else 'Sin datos'),
        ("Ciudad", escuela.ciudad.nombre if escuela.ciudad else 'Sin datos'),
        ("Predio", escuela.predio.numero_predio if escuela.predio else 'Sin datos'),
        ("Coordenadas", escuela.coordenadas),
        ("Latitud", escuela.latitud),
        ("Longitud", escuela.longitud),
        ("Tiene Internet", 'Sí' if escuela.tiene_internet else 'No'),
        ("Tiene Piso Tecnológico", 'Sí' if escuela.tiene_piso_tecnologico else 'No'),
    ]
    
    for campo, valor in datos_escuela:
        hoja[f'A{row_num}'] = campo
        hoja[f'B{row_num}'] = valor
        hoja[f'A{row_num}'].font = fuente_negrita
        row_num += 1
    
    row_num += 2

    # Bloque de Conectividad
    hoja.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
    hoja[f'A{row_num}'] = "DETALLES DE CONECTIVIDAD"
    hoja[f'A{row_num}'].font = fuente_negrita
    hoja[f'A{row_num}'].fill = relleno_cabecera
    hoja[f'A{row_num}'].alignment = Alignment(horizontal='center')
    row_num += 1

    headers_conectividad = ["Proveedor", "Estado", "Velocidad (Mbps)", "Método de Solicitud",  "Fecha de Instalación", "Fecha de Mejora", "Observaciones"]
    for col_num, header in enumerate(headers_conectividad, 1):
        cell = hoja.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = fuente_negrita
        cell.border = borde_fino
        cell.fill = relleno_cabecera
    row_num += 1

    if servicios.exists():
        for servicio in servicios:
            hoja.cell(row=row_num, column=1, value=servicio.proveedor.nombre if servicio.proveedor else 'Sin datos')
            hoja.cell(row=row_num, column=2, value=servicio.estado_conectividad.nombre if servicio.estado_conectividad else 'Sin datos')
            hoja.cell(row=row_num, column=3, value=servicio.velocidad_mbps)
            hoja.cell(row=row_num, column=4, value=servicio.metodo_solicitud.nombre if servicio.metodo_solicitud else 'Sin datos')
            hoja.cell(row=row_num, column=5, value=servicio.fecha_instalacion)
            hoja.cell(row=row_num, column=6, value=servicio.fecha_mejora)
            hoja.cell(row=row_num, column=7, value=servicio.observaciones)
            for col in range(1, 8):
                hoja.cell(row=row_num, column=col).border = borde_fino
            row_num += 1
    else:
        hoja.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
        hoja[f'A{row_num}'] = "No hay datos de conectividad."
        hoja[f'A{row_num}'].alignment = Alignment(horizontal='center')
        hoja[f'A{row_num}'].font = Font(italic=True)
        row_num += 1
    
    row_num += 2

    # Bloque de Piso Tecnológico
    hoja.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
    hoja[f'A{row_num}'] = "DETALLES DE PISO TECNOLÓGICO"
    hoja[f'A{row_num}'].font = fuente_negrita
    hoja[f'A{row_num}'].fill = relleno_cabecera
    hoja[f'A{row_num}'].alignment = Alignment(horizontal='center')
    row_num += 1

    headers_piso = ["Proveedor", "Tipo de Piso", "Plan de Piso", "Tipo de Mejora", "Fecha de Finalización", "Fecha de Mejora", "Observaciones"]
    for col_num, header in enumerate(headers_piso, 1):
        cell = hoja.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = fuente_negrita
        cell.border = borde_fino
        cell.fill = relleno_cabecera
    row_num += 1
    
    if pisos_tecnologicos.exists():
        for piso in pisos_tecnologicos:
            hoja.cell(row=row_num, column=1, value=piso.proveedor.nombre if piso.proveedor else 'Sin datos')
            hoja.cell(row=row_num, column=2, value=piso.tipo_piso_instalado.nombre if piso.tipo_piso_instalado else 'Sin datos')
            hoja.cell(row=row_num, column=3, value=piso.plan_piso.nombre if piso.plan_piso else 'Sin datos')
            hoja.cell(row=row_num, column=4, value=piso.tipo_mejora)
            hoja.cell(row=row_num, column=5, value=piso.fecha_terminado)
            hoja.cell(row=row_num, column=6, value=piso.fecha_mejora)
            hoja.cell(row=row_num, column=7, value=piso.observaciones)
            for col in range(1, 8):
                hoja.cell(row=row_num, column=col).border = borde_fino
            row_num += 1
    else:
        hoja.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
        hoja[f'A{row_num}'] = "No hay datos de piso tecnológico."
        hoja[f'A{row_num}'].alignment = Alignment(horizontal='center')
        hoja[f'A{row_num}'].font = Font(italic=True)
        row_num += 1
    
    # Ajustar ancho de columnas
    for col in range(1, hoja.max_column + 1):
        column_letter = get_column_letter(col)
        hoja.column_dimensions[column_letter].width = 25
    
    workbook.save(response)
    return response
# =========================================================================
# --- Genera el  reporte de filtro avanzado  en excel   ---
# =========================================================================
def exportar_resultados_excel(request):
    # 1. Reutiliza la lógica de filtrado de tu vista resultados_busqueda
    #    Aquí asumo que el filtrado inicial se hace en la vista de búsqueda
    
    # Crea un QueryDict con los parámetros de la URL
    filtros = request.GET
    queryset = Escuela.objects.all() # O el queryset base que uses

    # Aplica los filtros de la misma manera que en resultados_busqueda
    # EJEMPLOS DE FILTRADO (DEBES ADAPTAR ESTO A CÓMO ESTÁ FILTRANDO TU VISTA):
    if filtros.get('cue'):
        queryset = queryset.filter(cue__icontains=filtros.get('cue'))
    if filtros.get('region'):
        queryset = queryset.filter(region_id=filtros.get('region'))
    # ... continúa con todos tus filtros ...
    
    # 2. Creación del Libro de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="resultados_busqueda.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color='007bff', end_color='007bff', fill_type='solid') # Azul de Bootstrap
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # 3. Encabezados (Deben coincidir con los de tu tabla)
    headers = ['CUE', 'Nombre', 'Región', 'Distrito', 'Predio', 'Tiene Internet', 'Piso Tecnológico']
    ws.append(headers)

    # Aplicar estilos a la cabecera
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 4. Llenar filas con datos
    for escuela in queryset:
        row = [
            escuela.cue,
            escuela.nombre,
            escuela.region.nombre if escuela.region else 'N/A',
            escuela.distrito.nombre if escuela.distrito else 'N/A',
            escuela.predio.numero_predio if escuela.predio else 'N/A',
            'SÍ' if escuela.tiene_internet else 'NO',
            'SÍ' if escuela.tiene_piso_tecnologico else 'NO',
        ]
        ws.append(row)
    
    # 5. Ajustar ancho de columnas (Opcional, mejora visual)
    column_widths = [15, 60, 20, 20, 15, 15, 20] # Ejemplo de anchos
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # 6. Guardar y retornar
    wb.save(response)
    return response

# IMPORTANTE: Asegúrate de que la lógica de filtrado de esta vista 
# (punto 1) sea exactamente la misma que usas en tu vista resultados_busqueda.




from django.shortcuts import render
from django.db.models import Count
from .models import Escuela

# =========================================================================
# --- Vistas de Reportes (Nueva Vista para reportes_generales.html) ---
# =========================================================================

# ... (otras importaciones y código)

# =========================================================================
# --- Vistas de Reportes (Nueva Vista para reportes_generales.html) ---
# =========================================================================

# gestor/views.py

# ... otras funciones y código ...

from django.db.models import Count, Q
from .models import Escuela, Categoria # Importa tus modelos

def reportes_generales(request):
    """
    Genera el Reporte General de Cobertura, permitiendo filtrar los datos
    por Región y/o Distrito.
    """

    # --- 1. CAPTURA Y APLICACIÓN DE FILTROS AL QUERYSET BASE ---

    # Captura los IDs de filtro enviados por el formulario GET
    filtro_region_id = request.GET.get('region')
    filtro_distrito_id = request.GET.get('distrito')
    
    # Inicializa el QuerySet principal con todas las escuelas
    escuelas_queryset = Escuela.objects.all()
    
    # Define el título inicial
    titulo_pagina = 'Reportes Generales de Cobertura'
    
    # Aplica el filtro por Región
    if filtro_region_id:
        escuelas_queryset = escuelas_queryset.filter(region_id=filtro_region_id)
        try:
            region_nombre = Region.objects.get(id=filtro_region_id).nombre 
            titulo_pagina = f"Reporte por Región: {region_nombre}"
        except Region.DoesNotExist:
            pass # Si el ID no existe, usamos el título por defecto
    
    # Aplica el filtro por Distrito (se aplica después de Región si ambos están presentes)
    if filtro_distrito_id:
        escuelas_queryset = escuelas_queryset.filter(distrito_id=filtro_distrito_id)
        try:
            distrito_nombre = Distrito.objects.get(id=filtro_distrito_id).nombre
            titulo_pagina = f"Reporte por Distrito: {distrito_nombre}"
        except Distrito.DoesNotExist:
            pass # Si el ID no existe, usamos el título por defecto
    
    # -------------------------------------------------------------------------
    # 2. CÁLCULO DE TOTALES GLOBALES CON EL FILTRO APLICADO
    # -------------------------------------------------------------------------
    
    # Usamos el QuerySet filtrado (escuelas_queryset)
    total_escuelas = escuelas_queryset.count()
    
    # Conectividad (Internet)
    con_internet = escuelas_queryset.filter(tiene_internet=True).count()
    sin_internet = total_escuelas - con_internet
    
    # Equipamiento (Piso Tecnológico)
    con_piso = escuelas_queryset.filter(tiene_piso_tecnologico=True).count()
    sin_piso = total_escuelas - con_piso

    # -------------------------------------------------------------------------
    # 3. REPORTE DE CONECTIVIDAD (INTERNET) - DETALLE POR CATEGORÍA
    # -------------------------------------------------------------------------
    
    # Para el reporte por categoría, filtramos las Categorías por las escuelas
    # que pertenecen al QuerySet filtrado (escuelas_queryset).
    
    # NOTA: Para que esto funcione, necesitamos que la relación entre Categoria y Escuela
    # sea ManyToOne (Categoria tiene un set de 'escuela_set'). Si la relación es indirecta, 
    # se podría necesitar una consulta más compleja, pero esta estructura es la estándar.

    categoria_counts = []
    
    # Iteramos sobre TODAS las categorías disponibles
    for categoria in Categoria.objects.all().order_by('nombre'):
        # Filtramos nuestro QuerySet principal por la Categoría actual
        escuelas_en_categoria_y_filtro = escuelas_queryset.filter(categoria=categoria)
        
        total_cat = escuelas_en_categoria_y_filtro.count()
        
        # Solo procesamos si hay escuelas en esta categoría Y que cumplen el filtro
        if total_cat > 0:
            con_internet_cat = escuelas_en_categoria_y_filtro.filter(tiene_internet=True).count()
            porcentaje_cobertura = (float(con_internet_cat) / total_cat) * 100 
            
            categoria_counts.append({
                'nombre': categoria.nombre,
                'total_escuelas': total_cat,
                'con_internet': con_internet_cat, 
                'porcentaje_cobertura': round(porcentaje_cobertura, 1) 
            })

    # -------------------------------------------------------------------------
    # 4. REPORTE DE PISO TECNOLÓGICO - DETALLE POR CATEGORÍA
    # -------------------------------------------------------------------------
    
    categoria_piso_counts = []

    # Iteramos sobre TODAS las categorías disponibles
    for categoria in Categoria.objects.all().order_by('nombre'):
        # Filtramos nuestro QuerySet principal por la Categoría actual
        escuelas_en_categoria_y_filtro = escuelas_queryset.filter(categoria=categoria)
        
        total_cat = escuelas_en_categoria_y_filtro.count()
        
        # Solo procesamos si hay escuelas en esta categoría Y que cumplen el filtro
        if total_cat > 0:
            con_piso_cat = escuelas_en_categoria_y_filtro.filter(tiene_piso_tecnologico=True).count()
            porcentaje_cobertura_piso = (float(con_piso_cat) / total_cat) * 100 
            
            categoria_piso_counts.append({
                'nombre': categoria.nombre,
                'total_escuelas': total_cat,
                'con_piso': con_piso_cat,
                'porcentaje_cobertura_piso': round(porcentaje_cobertura_piso, 1)
            })

    # -------------------------------------------------------------------------
    # 5. DEFINICIÓN DEL CONTEXTO FINAL
    # -------------------------------------------------------------------------
    contexto = {
        # Para el formulario de filtro (listas desplegables y mantener la selección)
        'regiones': Region.objects.all().order_by('nombre'),
        'distritos': Distrito.objects.all().order_by('nombre'),
        'filtro_region_id': filtro_region_id,
        'filtro_distrito_id': filtro_distrito_id,
        
        'titulo_pagina': titulo_pagina,
        
        # Totales Globales (ahora filtrados)
        'total_escuelas': total_escuelas, 
        'con_internet': con_internet,
        'sin_internet': sin_internet,
        'con_piso': con_piso,
        'sin_piso': sin_piso,
        
        # Reportes por Categoría (ahora filtrados)
        'categoria_counts': categoria_counts,
        'categoria_piso_counts': categoria_piso_counts, 
    }
    
    return render(request, 'gestor/reportes_generales.html', contexto)

###Excel para hreporte general con filtro region y distrito

def exportar_reporte_excel(request):
    """
    Exporta el Reporte General de Cobertura a un archivo Excel, 
    aplicando los filtros de Región y Distrito si existen.
    """
    
    # 1. CAPTURA DE FILTROS (Igual que en reportes_generales)
    filtro_region_id = request.GET.get('region')
    filtro_distrito_id = request.GET.get('distrito')
    
    escuelas_queryset = Escuela.objects.all()
    nombre_archivo = "Reporte_Cobertura_General"
    
    if filtro_region_id:
        escuelas_queryset = escuelas_queryset.filter(region_id=filtro_region_id)
        try:
            region_nombre = Region.objects.get(id=filtro_region_id).nombre
            nombre_archivo = f"Reporte_Región_{region_nombre.replace(' ', '_')}"
        except Region.DoesNotExist:
            pass
            
    if filtro_distrito_id:
        escuelas_queryset = escuelas_queryset.filter(distrito_id=filtro_distrito_id)
        try:
            distrito_nombre = Distrito.objects.get(id=filtro_distrito_id).nombre
            nombre_archivo = f"Reporte_Distrito_{distrito_nombre.replace(' ', '_')}"
        except Distrito.DoesNotExist:
            pass

    # 2. CONFIGURACIÓN DE LA RESPUESTA HTTP PARA DESCARGA
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    # Define el nombre del archivo final
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}.xlsx'

    # Crea el libro de trabajo de Excel
    workbook = openpyxl.Workbook()
    
    
    # 3. HOJA 1: REPORTE DE CONECTIVIDAD (INTERNET)
    
    hoja_internet = workbook.active
    hoja_internet.title = "Internet por Categoria"
    
    # Encabezados de la tabla
    columnas_internet = [
        "Categoría", "Total Escuelas", "Escuelas con Internet", 
        "Escuelas sin Internet", "Cobertura Internet (%)"
    ]
    hoja_internet.append(columnas_internet)
    
    # Llenado de datos (usando la misma lógica de cálculo)
    categorias = Categoria.objects.all().order_by('nombre')
    
    for categoria in categorias:
        escuelas_cat = escuelas_queryset.filter(categoria=categoria)
        total_cat = escuelas_cat.count()
        
        if total_cat > 0:
            conectadas_cat = escuelas_cat.filter(tiene_internet=True).count()
            sin_internet_cat = total_cat - conectadas_cat
            porcentaje_cobertura = round((conectadas_cat / total_cat) * 100, 1)
            
            # Añadir fila a la hoja
            hoja_internet.append([
                categoria.nombre, 
                total_cat, 
                conectadas_cat, 
                sin_internet_cat,
                porcentaje_cobertura
            ])


    # 4. HOJA 2: REPORTE DE PISO TECNOLÓGICO
    
    hoja_piso = workbook.create_sheet(title="Piso Tecnologico por Categoria")
    
    # Encabezados de la tabla
    columnas_piso = [
        "Categoría", "Total Escuelas", "Escuelas con Piso Tecnológico", 
        "Escuelas sin Piso Tecnológico", "Cobertura Piso Tecnológico (%)"
    ]
    hoja_piso.append(columnas_piso)

    # Llenado de datos (usando la misma lógica de cálculo)
    for categoria in categorias: # Reutilizamos la lista de categorías
        escuelas_cat = escuelas_queryset.filter(categoria=categoria)
        total_cat = escuelas_cat.count()
        
        if total_cat > 0:
            con_piso_cat = escuelas_cat.filter(tiene_piso_tecnologico=True).count()
            sin_piso_cat = total_cat - con_piso_cat
            porcentaje_cobertura_piso = round((con_piso_cat / total_cat) * 100, 1)
            
            # Añadir fila a la hoja
            hoja_piso.append([
                categoria.nombre, 
                total_cat, 
                con_piso_cat, 
                sin_piso_cat,
                porcentaje_cobertura_piso
            ])
            

    # 5. GUARDAR Y DEVOLVER
    workbook.save(response)
    return response
##### FIN

    




    #---------------------------------------
    
### Para filtrar en mapa 
def mapa_escuelas_colores(request):
    regiones = Region.objects.all().order_by('nombre')
    distritos = Distrito.objects.all().order_by('nombre')
    estados = EstadoConectividad.objects.all().order_by('nombre')
    predios = Predio.objects.all().order_by('numero_predio')   # <-- agregado
    context = {
        'regiones': regiones,
        'distritos': distritos,
        'estados_conectividad': estados,
        'predios': predios,   # <-- agregado
    }
    return render(request, 'gestor/mapa_escuelas_colores.html', context)


def mapa_escuelas_con_internet(request):
    """Prepara datos para mostrar un mapa de solo las escuelas con Internet."""
    escuelas = Escuela.objects.filter(tiene_internet=True)
    
    escuelas_data = []
    for escuela in escuelas:
        if escuela.latitud and escuela.longitud:
            escuelas_data.append({
                'nombre': escuela.nombre,
                'latitud': float(escuela.latitud),
                'longitud': float(escuela.longitud),
                'cue': escuela.cue,
                'region_nombre': escuela.region.nombre if escuela.region else 'Sin datos'
            })
            
    context = {
        'escuelas_data': escuelas_data
    }
    
    return render(request, 'gestor/mapa_internet.html', context)

#Distritos 

def ajax_cargar_distritos(request):
    """
    Vista AJAX para cargar distritos basados en la selección de Regiones.
    """
    region_ids_str = request.GET.getlist('region_ids')
    
    try:
        # Filtra solo IDs que sean números y no estén vacíos
        region_ids = [int(id) for id in region_ids_str if id] 
    except ValueError:
        return JsonResponse({'opciones': {}}, status=200)

    distritos_queryset = []
    if region_ids:
        # Asumiendo que 'Distrito' tiene un campo de relación llamado 'region'
        distritos_queryset = Distrito.objects.filter(region__id__in=region_ids).order_by('nombre')

    opciones = {distrito.id: distrito.nombre for distrito in distritos_queryset}
    
    return JsonResponse({'opciones': opciones})

# Api par a que el mapa se vea  con los datos por sectores 

def api_escuelas_bounds(request):
    try:
        min_lat = float(request.GET.get('minLat'))
        max_lat = float(request.GET.get('maxLat'))
        min_lng = float(request.GET.get('minLng'))
        max_lng = float(request.GET.get('maxLng'))
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Parámetros de bounds inválidos'}, status=400)

    qs = Escuela.objects.filter(
        latitud__gte=min_lat,
        latitud__lte=max_lat,
        longitud__gte=min_lng,
        longitud__lte=max_lng,
    )

    # -------------------------
    # FILTROS OPCIONALES
    # -------------------------

    region_id = request.GET.get('region')
    distrito_id = request.GET.get('distrito')
    tiene_internet = request.GET.get('tiene_internet')
    tiene_piso = request.GET.get('tiene_piso')
    estado_id = request.GET.get('estado_conectividad')
    cue = request.GET.get('cue')
    predio = request.GET.get('predio')

    if region_id:
        qs = qs.filter(region__id=region_id)

    if distrito_id:
        qs = qs.filter(distrito__id=distrito_id)

    if tiene_internet in ('1', '0'):
        qs = qs.filter(tiene_internet=(tiene_internet == '1'))

    if tiene_piso in ('1', '0'):
        qs = qs.filter(tiene_piso_tecnologico=(tiene_piso == '1'))

    if estado_id:
        qs = qs.filter(servicioconectividad__estado_conectividad__id=estado_id).distinct()

    # -------------------------
    #  ✅ FILTRO POR CUE
    # -------------------------
    if cue:
        qs = qs.filter(cue__icontains=cue)

    # -------------------------
    # ✅ FILTRO POR PREDIO (FUNCIONANDO)
    #    Usa el campo correcto: numero_predio
    # -------------------------
    if predio:
        qs = qs.filter(predio__numero_predio__icontains=predio)

    # -------------------------
    # SERIALIZACIÓN
    # -------------------------
    resultados = []
    for e in qs:
        resultados.append({
            'cue': e.cue,
            'nombre': str(e.nombre),
            'latitud': float(e.latitud) if e.latitud is not None else None,
            'longitud': float(e.longitud) if e.longitud is not None else None,
            'tiene_internet': bool(e.tiene_internet),
            'tiene_piso_tecnologico': bool(e.tiene_piso_tecnologico),
            'region_id': e.region.id if e.region else None,
            'distrito_id': e.distrito.id if e.distrito else None,
        })

    return JsonResponse(resultados, safe=False)



def api_escuela(request, cue):
    """
    Devuelve detalles serializables de una escuela para el popup.
    Evita devolver objetos ORM; siempre strings o None.
    """
    escuela = get_object_or_404(Escuela, cue=cue)

    # Obtenemos primer servicio (podés ajustar: último por fecha, etc.)
    servicio = ServicioConectividad.objects.filter(escuela=escuela).first()
    piso = PisoTecnologico.objects.filter(escuela=escuela).first()

    data = {
        'cue': escuela.cue,
        'nombre': escuela.nombre,
        'categoria': escuela.categoria.nombre if escuela.categoria else None,
        'latitud': float(escuela.latitud) if escuela.latitud is not None else None,
        'longitud': float(escuela.longitud) if escuela.longitud is not None else None,
        # Servicio
        'estado_conectividad': servicio.estado_conectividad.nombre if (servicio and servicio.estado_conectividad) else None,
        'proveedor': servicio.proveedor.nombre if (servicio and servicio.proveedor) else None,
        'velocidad_mbps': servicio.velocidad_mbps if servicio else None,
        # Piso
        'plan_piso': piso.plan_piso.nombre if (piso and piso.plan_piso) else None,
        'tipo_piso': piso.tipo_piso_instalado.nombre if (piso and piso.tipo_piso_instalado) else None,
        'proveedor_piso': piso.proveedor.nombre if (piso and piso.proveedor) else None,
        'observaciones_servicio': servicio.observaciones if servicio else None,
        'observaciones_piso': piso.observaciones if piso else None,
    }

    return JsonResponse(data)